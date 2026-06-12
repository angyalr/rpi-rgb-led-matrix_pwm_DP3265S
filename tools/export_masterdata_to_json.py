import argparse
import json
from datetime import date, datetime

import openpyxl


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value != "" else None
    if isinstance(value, (datetime, date)):
        return value.isoformat(sep=" ")
    return value


def to_int(value):
    value = clean(value)
    if value is None:
        return None
    return int(float(value))


def to_float(value):
    value = clean(value)
    if value is None:
        return None
    return float(value)


def to_datetime_text(value):
    value = clean(value)
    if value is None:
        return None
    if isinstance(value, str):
        for fmt in ("%Y. %m. %d. %H:%M:%S", "%Y.%m.%d. %H:%M:%S"):
            try:
                return datetime.strptime(value, fmt).strftime("%Y-%m-%d %H:%M:%S")
            except ValueError:
                pass
    return value


def row_dict(headers, row):
    return {headers[i]: clean(row[i]) if i < len(row) else None for i in range(len(headers))}


def load_table(ws, header_row=1):
    headers = [clean(cell.value) for cell in ws[header_row]]
    rows = []
    for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(clean(v) is not None for v in values):
            continue
        rows.append(row_dict(headers, values))
    return rows


def export_rendszam(ws):
    rows = []
    seen = set()
    # The sheet stores one vehicle over two physical rows: data row, then country row.
    for row in range(5, ws.max_row + 1, 2):
        rendszam = clean(ws.cell(row, 1).value)
        potkocsi = clean(ws.cell(row, 7).value)
        if not rendszam and not potkocsi:
            continue
        key = ((rendszam or "").upper(), (potkocsi or "").upper())
        if key in seen:
            continue
        seen.add(key)
        country = clean(ws.cell(row + 1, 4).value) if row + 1 <= ws.max_row else None
        pcountry = clean(ws.cell(row + 1, 5).value) if row + 1 <= ws.max_row else None
        rows.append(
            {
                "Rendszam": rendszam or "",
                "PotKocsi": potkocsi or "",
                "Country": country,
                "PCountry": pcountry,
            }
        )
    return rows


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("excel")
    parser.add_argument("output")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.excel, data_only=True)

    vevok = []
    for r in load_table(wb["Ügyfél"]):
        vevok.append(
            {
                "Vevokod": to_int(r.get("Ügyfélkód")),
                "Vevonev": clean(r.get("Ügyfélnév")),
                "Adoszam": clean(r.get("Adószám")),
                "Country": clean(r.get("Ország")),
                "VevoIrSzam": clean(r.get("Ir.szám")),
                "VevoVaros": clean(r.get("Város")),
                "VevoUtca": clean(r.get("Utca")),
                "StreetType": clean(r.get("StreetType")),
                "VevoHsz": clean(r.get("Házszám")),
                "LotNumber": clean(r.get("LotNumber")),
                "BankszamlaSzam": clean(r.get("Bankszámla szám")),
                "CegjegyzekSzam": clean(r.get("Cégjegyzék szám")),
                "CegjegyzesreJogosult": clean(r.get("Cégjegyzésre jogosult")),
                "CegjegyzoJogosultsaga": clean(r.get("Cégjegyző jogosultsága")),
                "SAP": clean(r.get("SAP")),
                "Hitel": to_float(r.get("Hitel limit")),
                "Contact": clean(r.get("Kapcsolat tartó")),
                "Mobil": clean(r.get("Mobil")),
                "email": clean(r.get("email")),
                "TimeFormat": clean(r.get("TimeFormat")),
                "ModifyInDate": to_int(r.get("Javítás adatokban hó")),
                "KshId": clean(r.get("KshId") or r.get("FELIR azonosító")),
                "EKAERStatus": clean(r.get("EKAERStatus")),
                "Szerzodes": to_int(r.get("Árlista")),
            }
        )

    telep = []
    for r in load_table(wb["Lerakodási lista"]):
        telep.append(
            {
                "VevoKod": to_int(r.get("VevoKod")),
                "TelepKod": to_int(r.get("Lerakodásih.kód")),
                "TelepHely": clean(r.get("Lerakodásih.név")),
                "Country": clean(r.get("Ország")),
                "ZipCode": clean(r.get("Ir.szám")),
                "City": clean(r.get("Település")),
                "Street": clean(r.get("Közterület")),
                "StreetType": clean(r.get("Közterület típusa")),
                "StreetNumber": clean(r.get("Házszám")),
                "LotNumber": clean(r.get("Helyrajzi szám")),
                "Email": clean(r.get("Email")),
                "Phone": clean(r.get("Telefon")),
                "Contact": clean(r.get("Kapcsolat tartó")),
                "VATNumber": clean(r.get("Adószám")),
            }
        )

    aruk = []
    for r in load_table(wb["Áru"]):
        aruk.append(
            {
                "Arukod": to_int(r.get("Árukód")),
                "Arunev": clean(r.get("Árunév")),
                "Egysegar": to_float(r.get("Egységár")),
                "MEgyseg": clean(r.get("M.egység")),
                "MerlegValtoSzam": to_float(r.get("Váltószám")),
                "Modositva": to_datetime_text(r.get("Modositva")),
                "VatExemptionCase": clean(r.get("VatExemptionCase")),
                "VatExemptionReason": clean(r.get("VatExemptionReason")),
                "VTSZ": clean(r.get("VTSZ")),
                "Afakulcs": to_float(r.get("Áfakulcs")),
                "adrNumber": clean(r.get("veszélyes besorolás")),
            }
        )

    ar = []
    for r in load_table(wb["Különárak"]):
        ar.append(
            {
                "Szerzodes": to_int(r.get("Árlista")),
                "TelepKod": to_int(r.get("TelepKod")),
                "Arukod": to_int(r.get("Árukód")),
                "Egysegar": to_float(r.get("Egységár")),
            }
        )

    data = {
        "VEVOK": vevok,
        "Telep": telep,
        "ARUK": aruk,
        "Ar": ar,
        "RENDSZAM": export_rendszam(wb["Rendszám"]),
    }

    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(json.dumps({k: len(v) for k, v in data.items()}, ensure_ascii=False))


if __name__ == "__main__":
    main()
